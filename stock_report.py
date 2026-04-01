# ╔══════════════════════════════════════════════════════════════════╗
# ║        台股日報 × Resend 自動寄信 — Google Colab 版             ║
# ║  將每個 # ── CELL N 區塊分別貼入 Colab cell，依序執行           ║
# ║  Cell 1–4 只需執行一次；Cell 5（防斷線）+ Cell 6（排程）        ║
# ║  依序執行後即可每天台灣時間 14:00 自動寄出報告                  ║
# ╚══════════════════════════════════════════════════════════════════╝


# ── CELL 1：安裝套件 ──────────────────────────────────────────────
# !pip install -q yfinance openpyxl pandas pytz matplotlib numpy Pillow resend


# ── CELL 2：設定（只需修改這裡）─────────────────────────────────
RESEND_API_KEY = "re_xxxxxxxxxxxxxx"       # ← 你的 Resend API Key
EMAIL_FROM     = "reports@yourdomain.com"  # ← Resend 已驗證的寄件地址
EMAIL_TO       = ["you@example.com"]       # ← 收件人（可多位）
EMAIL_SUBJECT  = "【台股日報】每日收盤股價報告"

SHEET_ID = "1bg8oZiiDFLMisYBXaDF20aX5MPxjLRNac81I_kWoQGQ"
GID      = "1668869607"


# ── CELL 3：import + 共用樣式 ─────────────────────────────────────
import io, time, base64, logging
from datetime import datetime, timedelta

import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
import matplotlib.dates as mdates
from matplotlib.ticker import MaxNLocator
import matplotlib.patheffects as pe
import numpy as np
import pandas as pd
import pytz
import resend
import yfinance as yf
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

logging.basicConfig(level=logging.INFO, format="%(asctime)s %(levelname)s %(message)s")
resend.api_key = RESEND_API_KEY

TW_TZ = pytz.timezone("Asia/Taipei")

def mk_fill(c): return PatternFill("solid", start_color=c)
def mk_border(c="BFBFBF"):
    s = Side(style="thin", color=c)
    return Border(left=s, right=s, top=s, bottom=s)

C_DARK="1F4E79"; C_MID="2E75B6"; C_LIGHT="EBF3FB"; C_WHITE="FFFFFF"
C_RED_BG="FFDCDC"; C_RED_FG="C00000"; C_GRN_BG="D6F0D6"; C_GRN_FG="375623"; C_GREY="7F7F7F"
THIN=mk_border()
CENTER=Alignment(horizontal="center", vertical="center")
RIGHT=Alignment(horizontal="right", vertical="center")
def hfont(sz=10): return Font(name="Arial", bold=True, color=C_WHITE, size=sz)
def nfont(sz=10): return Font(name="Arial", size=sz)

print("✅ 套件與設定載入完成")


# ── CELL 4：核心函式 ──────────────────────────────────────────────

def fetch(ticker, start, end):
    try:
        raw = yf.download(ticker,
                          start=start.strftime("%Y-%m-%d"),
                          end=(end + timedelta(days=1)).strftime("%Y-%m-%d"),
                          interval="1d", auto_adjust=True, progress=False)
        if raw.empty: return None
        df = raw[["Open","High","Low","Close","Volume"]].copy()
        if isinstance(df.columns, pd.MultiIndex):
            df.columns = df.columns.get_level_values(0)
        df.columns = ["開盤價","最高價","最低價","收盤價(13:30)","成交量"]
        idx = pd.to_datetime(df.index)
        idx = idx.tz_localize("UTC").tz_convert(TW_TZ) if idx.tz is None else idx.tz_convert(TW_TZ)
        df.index = idx.strftime("%Y-%m-%d"); df.index.name = "日期"
        df = df.sort_index()
        df["漲跌"]      = df["收盤價(13:30)"].diff().round(2)
        df["漲跌幅(%)"] = (df["收盤價(13:30)"].pct_change() * 100).round(2)
        return df
    except Exception as e:
        logging.warning(f"✗ {ticker}: {e}"); return None


def make_chart(ticker, df, names):
    close = df["收盤價(13:30)"].dropna()
    dates = pd.to_datetime(close.index)
    prices = close.values.astype(float)
    x_num = np.arange(len(prices))
    coeffs = np.polyfit(x_num, prices, 1)
    trend  = np.polyval(coeffs, x_num)
    is_up  = trend[-1] >= trend[0]

    lc = "#E53E3E" if is_up else "#38A169"
    tc = "#F6AD55" if is_up else "#63B3ED"
    bg = "#0D1117"; gd = "#2D3748"; tx = "#E2E8F0"; lb = "#A0AEC0"

    fig, ax = plt.subplots(figsize=(13, 4.8), facecolor=bg)
    ax.set_facecolor(bg)
    ax.fill_between(dates, prices, prices.min()*0.998, color=lc, alpha=0.12, zorder=1)
    ax.plot(dates, prices, color=lc, linewidth=1.8, zorder=3, label="收盤價(13:30)",
            path_effects=[pe.Stroke(linewidth=3, foreground=bg), pe.Normal()])
    sp = (trend[-1]-trend[0])/trend[0]*100
    ax.plot(dates, trend, color=tc, linewidth=1.6, linestyle="--", zorder=4,
            label=f"趨勢線  {sp:+.2f}%", alpha=0.9)

    for ip, fmt, va, clr in [
        (np.argmax(prices), f"▲ {prices[np.argmax(prices)]:,.2f}", "bottom", "#FC8181"),
        (np.argmin(prices), f"▼ {prices[np.argmin(prices)]:,.2f}", "top",    "#68D391"),
    ]:
        ax.annotate(fmt, xy=(dates[ip], prices[ip]),
                    xytext=(0, 10 if va=="bottom" else -10), textcoords="offset points",
                    ha="center", va=va, color=clr, fontsize=8, fontweight="bold",
                    arrowprops=dict(arrowstyle="-", color=clr, lw=0.8))

    ax.xaxis.set_major_formatter(mdates.DateFormatter("%m/%d"))
    ax.xaxis.set_major_locator(mdates.WeekdayLocator(byweekday=mdates.MO, interval=2))
    plt.setp(ax.xaxis.get_majorticklabels(), rotation=0, color=lb, fontsize=8)
    ax.yaxis.set_major_locator(MaxNLocator(nbins=6))
    plt.setp(ax.yaxis.get_majorticklabels(), color=lb, fontsize=8)
    ax.tick_params(axis="both", which="both", length=0, colors=lb)
    ax.spines[:].set_visible(False)
    ax.grid(axis="y", color=gd, linewidth=0.6)
    ax.grid(axis="x", color=gd, linewidth=0.4, linestyle=":")
    ax.set_axisbelow(True)

    last_p = prices[-1]; first_p = prices[0]
    chg = (last_p-first_p)/first_p*100
    cc = "#FC8181" if chg >= 0 else "#68D391"
    cs = "▲" if chg >= 0 else "▼"
    company = names.get(ticker, "")
    title = f"{company}（{ticker}）  收盤價走勢圖（台灣時間 13:30）" if company else f"{ticker}  收盤價走勢圖（台灣時間 13:30）"
    fig.text(0.015, 0.97, title, ha="left", va="top", color=tx, fontsize=11, fontweight="bold", transform=fig.transFigure)
    fig.text(0.015, 0.83, f"最新收盤 {last_p:,.2f}  {cs} {abs(chg):.2f}%（近三個月）",
             ha="left", va="top", color=cc, fontsize=9, transform=fig.transFigure)
    ax.legend(loc="upper left", frameon=True, framealpha=0.3, facecolor=bg,
              edgecolor=gd, labelcolor=tx, fontsize=8.5, handlelength=2)
    ax.set_xlim(dates[0], dates[-1])
    fig.text(0.99, 0.02,
             f"{dates[0].strftime('%Y/%m/%d')}  →  {dates[-1].strftime('%Y/%m/%d')}  │  共 {len(prices)} 個交易日",
             ha="right", va="bottom", color=lb, fontsize=7.5, transform=fig.transFigure)
    fig.tight_layout(rect=[0, 0.04, 1, 0.90])
    buf = io.BytesIO()
    fig.savefig(buf, format="png", dpi=130, bbox_inches="tight", facecolor=bg)
    plt.close(fig); buf.seek(0)
    return buf.read()


def write_stock_sheet(wb, ticker, df, names, start, end):
    safe = ticker.replace(".TWO","^").replace(".TW","")
    ws   = wb.create_sheet(title=safe)
    company = names.get(ticker,"")
    dn = f"{company}（{ticker}）" if company else ticker

    ws.merge_cells("A1:H1")
    tc = ws["A1"]
    tc.value = f"{dn}  每日 13:30 收盤股價｜{start.strftime('%Y/%m/%d')} – {end.strftime('%Y/%m/%d')}"
    tc.font=Font(name="Arial",bold=True,color=C_WHITE,size=12); tc.fill=mk_fill(C_DARK); tc.alignment=CENTER
    ws.row_dimensions[1].height=26

    hdrs=["日期","開盤價","最高價","最低價","收盤價(13:30)","成交量","漲跌","漲跌幅(%)"]
    for ci,h in enumerate(hdrs,1):
        c=ws.cell(row=2,column=ci,value=h)
        c.font=hfont(); c.fill=mk_fill(C_MID); c.alignment=CENTER; c.border=THIN
    ws.row_dimensions[2].height=18

    for ri,(dt,row) in enumerate(df.iterrows(),1):
        er=ri+2; bf=mk_fill(C_LIGHT) if ri%2==0 else mk_fill(C_WHITE)
        vals=[dt,row["開盤價"],row["最高價"],row["最低價"],row["收盤價(13:30)"],row["成交量"],row["漲跌"],row["漲跌幅(%)"]]
        for ci,val in enumerate(vals,1):
            c=ws.cell(row=er,column=ci,value=val); c.border=THIN; c.font=nfont()
            if ci==1:   c.alignment=CENTER; c.fill=bf
            elif ci in (2,3,4,5,7): c.number_format="#,##0.00"; c.alignment=RIGHT; c.fill=bf
            elif ci==6: c.number_format="#,##0"; c.alignment=RIGHT; c.fill=bf
            elif ci==8:
                c.number_format="+0.00;-0.00;0.00"; c.alignment=RIGHT
                try:    fv=float(val) if val is not None and str(val)!="nan" else 0
                except: fv=0
                if fv>0:   c.fill=mk_fill(C_RED_BG); c.font=Font(name="Arial",size=10,color=C_RED_FG)
                elif fv<0: c.fill=mk_fill(C_GRN_BG); c.font=Font(name="Arial",size=10,color=C_GRN_FG)
                else:      c.fill=bf

    for ci,w in enumerate([13,10,10,10,15,14,9,11],1):
        ws.column_dimensions[get_column_letter(ci)].width=w
    ws.freeze_panes="A3"

    cs = df.shape[0]+5
    img=XLImage(io.BytesIO(make_chart(ticker,df,names))); img.width=900; img.height=340
    ws.add_image(img,f"A{cs}")
    ws.merge_cells(f"A{cs-1}:H{cs-1}")
    lc=ws.cell(row=cs-1,column=1,value="▼  收盤價走勢圖 + 趨勢線（虛線）")
    lc.font=Font(name="Arial",bold=True,color=C_WHITE,size=10); lc.fill=mk_fill(C_DARK); lc.alignment=CENTER
    ws.row_dimensions[cs-1].height=20
    nr=cs+27
    nc=ws.cell(row=nr,column=1,value=f"※ 資料來源：Yahoo Finance｜下載時間：{datetime.now(TW_TZ).strftime('%Y-%m-%d %H:%M')} 台灣時間")
    nc.font=Font(name="Arial",size=8,color=C_GREY,italic=True)
    ws.merge_cells(f"A{nr}:H{nr}")


def write_summary(wb, rows, start, end):
    ws=wb.create_sheet(title="📋 匯總總覽",index=0)
    ws.merge_cells("A1:I1")
    tc=ws["A1"]
    tc.value=f"台股多標的收盤價匯總｜{start.strftime('%Y/%m/%d')} – {end.strftime('%Y/%m/%d')}"
    tc.font=Font(name="Arial",bold=True,color=C_WHITE,size=13); tc.fill=mk_fill(C_DARK); tc.alignment=CENTER
    ws.row_dimensions[1].height=28
    hdrs=["股票代碼","市場","筆數","期間起始收盤","最新收盤(13:30)","期間最高","期間最低","期間漲跌幅(%)","備註"]
    for ci,h in enumerate(hdrs,1):
        c=ws.cell(row=2,column=ci,value=h)
        c.font=hfont(); c.fill=mk_fill(C_MID); c.alignment=CENTER; c.border=THIN
    ws.row_dimensions[2].height=18
    for ri,row in enumerate(rows,1):
        er=ri+2; bf=mk_fill(C_LIGHT) if ri%2==0 else mk_fill(C_WHITE)
        for ci,val in enumerate(row,1):
            c=ws.cell(row=er,column=ci,value=val); c.border=THIN; c.font=nfont()
            if ci in (1,2,9): c.alignment=CENTER; c.fill=bf
            elif ci==3:       c.number_format="0"; c.alignment=CENTER; c.fill=bf
            elif ci in (4,5,6,7): c.number_format="#,##0.00"; c.alignment=RIGHT; c.fill=bf
            elif ci==8:
                c.number_format="+0.00;-0.00;0.00"; c.alignment=RIGHT
                try:    fv=float(val) if val is not None and str(val)!="nan" else 0
                except: fv=0
                if fv>0:   c.fill=mk_fill(C_RED_BG); c.font=Font(name="Arial",size=10,color=C_RED_FG)
                elif fv<0: c.fill=mk_fill(C_GRN_BG); c.font=Font(name="Arial",size=10,color=C_GRN_FG)
                else:      c.fill=bf
    for ci,w in enumerate([14,8,8,16,18,12,12,16,12],1):
        ws.column_dimensions[get_column_letter(ci)].width=w
    ws.freeze_panes="A3"
    nr=len(rows)+4
    nc=ws.cell(row=nr,column=1,
               value="※ 收盤價(13:30)：台灣證交所/櫃買每日13:30收盤｜資料來源：Yahoo Finance"
                     f"｜下載時間：{datetime.now(TW_TZ).strftime('%Y-%m-%d %H:%M')} 台灣時間")
    nc.font=Font(name="Arial",size=8,color=C_GREY,italic=True)
    ws.merge_cells(f"A{nr}:I{nr}")


print("✅ 核心函式載入完成")


# ── CELL 5：generate_report + send_email 函式定義 ─────────────────

def generate_report():
    now        = datetime.now(TW_TZ)
    end_date   = now.replace(hour=0, minute=0, second=0, microsecond=0)
    start_date = end_date - timedelta(days=91)

    url      = f"https://docs.google.com/spreadsheets/d/{SHEET_ID}/export?format=csv&gid={GID}"
    df_meta  = pd.read_csv(url)
    tickers  = df_meta["TICKERS"].tolist()
    names    = dict(zip(df_meta["TICKERS"], df_meta["公司名稱"])) if "公司名稱" in df_meta.columns else {}

    wb = Workbook(); wb.remove(wb.active)
    summary, ok = [], 0
    print(f"📥 開始下載 {len(tickers)} 支股票資料...\n")

    for i, ticker in enumerate(tickers, 1):
        market  = "上櫃(OTC)" if ticker.endswith(".TWO") else "上市(TWSE)"
        company = names.get(ticker, "")
        dn      = f"{company}（{ticker}）" if company else ticker
        print(f"  [{i:02d}/{len(tickers)}] {dn} ({market}) ...", end=" ", flush=True)
        df = fetch(ticker, start_date, end_date)
        if df is not None and not df.empty:
            write_stock_sheet(wb, ticker, df, names, start_date, end_date)
            cl  = df["收盤價(13:30)"].dropna()
            f0, fl = float(cl.iloc[0]), float(cl.iloc[-1])
            chg = round((fl-f0)/f0*100, 2) if f0 else None
            summary.append([ticker, market, len(df), f0, fl,
                             float(df["最高價"].max()), float(df["最低價"].min()), chg, "OK"])
            print(f"✅ {len(df)} 筆"); ok += 1
        else:
            summary.append([ticker, market, 0, None, None, None, None, None, "無資料"])
            print("⚠  跳過")
        time.sleep(0.3)

    write_summary(wb, summary, start_date, end_date)
    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    print(f"\n🎉 報告產生完成！成功 {ok}/{len(tickers)} 支")
    return buf.read()

def send_email(xlsx_bytes):
    now_str = datetime.now(TW_TZ).strftime("%Y-%m-%d")
    fname   = f"台股收盤股價_{now_str}.xlsx"
    params: resend.Emails.SendParams = {
        "from":    EMAIL_FROM,
        "to":      EMAIL_TO,
        "subject": f"{EMAIL_SUBJECT}｜{now_str}",
        "html": f"""
            <h2>台股每日收盤股價報告</h2>
            <p>日期：{now_str}</p>
            <p>請見附件 Excel 報告，內含各股走勢圖與匯總表。</p>
            <hr><small>資料來源：Yahoo Finance｜台灣時間 13:30 收盤價</small>
        """,
        "attachments": [{"filename": fname, "content": base64.b64encode(xlsx_bytes).decode()}],
    }
    result = resend.Emails.send(params)
    print(f"✅ 郵件已送出！ID：{result['id']}")

print("✅ generate_report / send_email 函式載入完成")


# ── CELL 5：防 Colab 閒置斷線（執行後保持連線）──────────────────
# 在瀏覽器開發人員工具的 Console 貼入以下 JavaScript，
# 或直接執行本 cell（會透過 IPython 注入 JS）。
# 作用：每 60 秒自動點擊 Colab 的連線按鈕，防止 90 分鐘閒置斷線。

from IPython.display import Javascript, display

display(Javascript("""
function keepAlive() {
  console.log('[KeepAlive] ' + new Date().toLocaleTimeString());
  // 嘗試點擊「重新連線」按鈕（若存在）
  const btn = document.querySelector('colab-connect-button');
  if (btn) btn.click();
}
// 每 55 秒執行一次（略低於 60 秒以確保觸發）
const _kaTimer = setInterval(keepAlive, 55000);
console.log('[KeepAlive] 已啟動，每 55 秒觸發一次');
"""))

print("✅ 防斷線 JS 已注入，請保持瀏覽器視窗開啟（不需在前景，但不能關閉分頁）")


# ── CELL 6：每天台灣時間 14:00 自動執行排程 ──────────────────────
# ⚡ 直接執行此 cell，程式會在背景等待並於每天 14:00（台灣時間）
#    自動產生報告並寄出，執行後 cell 會持續跑直到你手動中斷。
#
# 🔑 關鍵設計：
#   - 用台灣時間判斷，不依賴 Colab 系統時區（Colab 預設為 UTC）
#   - 記錄 last_run_date 避免同一天重複觸發
#   - 每 30 秒檢查一次時間，不佔用大量資源

SEND_HOUR   = 14   # 台灣時間幾點送出（24 小時制）
SEND_MINUTE = 0    # 幾分送出

last_run_date = None

print(f"⏰ 排程啟動！將於每天台灣時間 {SEND_HOUR:02d}:{SEND_MINUTE:02d} 自動寄出報告")
print(f"   目前台灣時間：{datetime.now(TW_TZ).strftime('%Y-%m-%d %H:%M:%S')}")
print("   （手動中斷請按左側 ■ 停止按鈕）\n")

while True:
    now_tw   = datetime.now(TW_TZ)
    today    = now_tw.date()
    is_time  = (now_tw.hour == SEND_HOUR and now_tw.minute == SEND_MINUTE)
    not_sent = (last_run_date != today)

    if is_time and not_sent:
        print(f"\n{'='*55}")
        print(f"🚀 [{now_tw.strftime('%Y-%m-%d %H:%M')}] 開始執行每日報告任務...")
        print(f"{'='*55}")
        try:
            xlsx = generate_report()
            send_email(xlsx)
            last_run_date = today
            print(f"✅ 今日任務完成｜下次執行：明天 {SEND_HOUR:02d}:{SEND_MINUTE:02d}\n")
        except Exception as e:
            print(f"❌ 任務失敗：{e}\n")
            # 失敗不記錄 last_run_date，60 秒後會重試（當分鐘內）
    else:
        # 每 10 分鐘印一次心跳，確認程式仍在運行
        if now_tw.minute % 10 == 0 and now_tw.second < 30:
            next_run = now_tw.replace(hour=SEND_HOUR, minute=SEND_MINUTE, second=0, microsecond=0)
            if next_run <= now_tw:
                next_run += timedelta(days=1)
            diff = next_run - now_tw
            h, m = divmod(int(diff.total_seconds()) // 60, 60)
            print(f"💓 [{now_tw.strftime('%H:%M')}] 等待中... 距下次執行還有 {h}h {m}m")

    time.sleep(30)
